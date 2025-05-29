
import streamlit as st
from pdf2image import convert_from_path
import pytesseract
import pandas as pd
import re
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font

st.set_page_config(page_title="Conversor PDF para Excel", layout="wide")
st.title("Conversor de PDF (Requisição) para Excel")
st.markdown("**Transforme arquivos PDF com itens de requisições em planilhas Excel formatadas.**")

# Upload do PDF
pdf_file = st.file_uploader("Envie o arquivo PDF", type=["pdf"])

if pdf_file:
    with open("temp.pdf", "wb") as f:
        f.write(pdf_file.read())

    pages = convert_from_path("temp.pdf", dpi=300)

    blocos = []
    bloco = ""
    for page in pages:
        texto = pytesseract.image_to_string(page, lang='por')
        for linha in texto.split('\n'):
            linha = linha.strip()
            if linha:
                bloco += linha + "\n"
                if "Valor total" in linha:
                    blocos.append(bloco)
                    bloco = ""

    def ajustar_texto(texto):
        texto = texto.strip().strip('"').strip("'")
        frases = re.split(r'(?<=[.!?])\s+', texto)
        frases_corrigidas = []
        for frase in frases:
            frase = frase.strip().strip('"').strip("'")
            if not frase:
                continue
            frase = frase.lower()
            primeira_letra = re.search(r"[A-Za-zÁÉÍÓÚáéíóúÀàÂâÊêÔôÇç]", frase)
            if primeira_letra:
                idx = primeira_letra.start()
                frase = frase[:idx] + frase[idx].upper() + frase[idx+1:]
            if frase[-1] not in ".!?":
                frase += "."
            frases_corrigidas.append(frase)
        return " ".join(frases_corrigidas)

    def extrair_unidade_e_quantidade(linhas):
        unidade = ""
        quantidade = ""
        for i, linha in enumerate(linhas):
            if "Unidade:" in linha:
                m = re.search(r"Unidade:\s*([A-Za-zçÇ ]+)", linha)
                if m:
                    unidade = m.group(1).strip().title()
                for j in range(1, 4):
                    if i + j < len(linhas):
                        prox = linhas[i + j].strip()
                        if "Quantidade:" in prox:
                            m2 = re.search(r"Quantidade:\s*(\d+)", prox)
                            if m2:
                                quantidade = int(m2.group(1))
                                return unidade, quantidade
                        elif re.fullmatch(r"\d{1,5}", prox):
                            quantidade = int(prox)
                            return unidade, quantidade
                break
        return unidade, quantidade

    def extrair_campos(texto, item_index):
        def campo(regex):
            r = re.search(regex, texto, re.DOTALL | re.IGNORECASE)
            return r.group(1).strip().replace('\n', ' ') if r else ""
        linhas = texto.split("\n")
        unidade, qtd = extrair_unidade_e_quantidade(linhas)
        descricao = campo(r"Descrição detalhada\s*:?.*?(.*?)Unidade")
        descricao = ajustar_texto(descricao)
        valor_unitario = campo(r"Valor unitário\s*:?.*?R\$\s*([\d.,]+)").replace(".", "").replace(",", ".")
        valor_total = campo(r"Valor total\s*:?.*?R\$\s*([\d.,]+)").replace(".", "").replace(",", ".")
        try: vu = float(valor_unitario)
        except: vu = None
        try: vt = float(valor_total)
        except: vt = None
        item = campo(r"Item\s*:?.*?(\d+)")
        item = int(item) if item.isdigit() else item_index + 1
        return {
            'ITEM': item,
            'CATMAT': int(campo(r"CATMAT\s*:?.*?(\d+)") or 0),
            'DESCRIÇÃO DETALHADA': descricao,
            'UNIDADE': unidade,
            'QUANTIDADE': qtd,
            'VALOR UNITÁRIO': vu,
            'VALOR TOTAL': vt,
        }

    dados = [extrair_campos(bloco, idx) for idx, bloco in enumerate(blocos)]
    df_final = pd.DataFrame(dados)

    wb = Workbook()
    ws = wb.active
    ws.title = "ITENS"

    for i, row in enumerate(dataframe_to_rows(df_final, index=False, header=True)):
        ws.append(row)

    moeda_fmt = '"R$"#,##0.00'
    for cell in ws[1]:
        cell.value = str(cell.value).upper()
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].alignment = row[1].alignment = row[4].alignment = Alignment(horizontal="center")
        row[2].alignment = Alignment(horizontal="justify", vertical="top", wrap_text=False)
        row[5].number_format = row[6].number_format = moeda_fmt
        row[5].alignment = row[6].alignment = Alignment(horizontal="center")

    from tempfile import NamedTemporaryFile
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        st.success("Planilha gerada com sucesso!")
        st.download_button("📥 Baixar Excel", data=open(tmp.name, "rb").read(), file_name="itens_extraidos.xlsx")
